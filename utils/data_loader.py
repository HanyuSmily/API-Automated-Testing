#这个文件负责责解析外部数据文件并将其参数化传入测试用例
import csv
import json
import os
from typing import Any, Dict, List, Text

from loguru import logger


class DataLoader:
    """数据加载器：解析 CSV / JSON / Excel 外部数据文件，输出参数化数据列表"""

    # 支持的文件扩展名 -> 对应解析方法
    _LOADERS = {
        ".csv": "_load_csv",
        ".json": "_load_json",
        ".xlsx": "_load_excel",
        ".xls": "_load_excel",
    }

    @classmethod
    def load(cls, file_path: Text) -> List[Dict[Text, Any]]:
        """统一入口：按扩展名自动选择解析器

        Args:
            file_path: 数据文件路径

        Returns:
            List[Dict]: 每行数据转换成一个字典（键为首行表头）
        """
        # 校验文件是否存在
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"数据文件不存在: {file_path}")

        # 取扩展名并小写
        ext = os.path.splitext(file_path)[1].lower()
        # 取对应解析方法名
        loader_name = cls._LOADERS.get(ext)
        if not loader_name:
            raise ValueError(f"不支持的数据文件类型: {ext}（支持: {list(cls._LOADERS.keys())}）")

        # 反射获取方法并调用变成新发文件数据
        loader = getattr(cls, loader_name)
        data = loader(file_path)

        logger.info(f"加载数据文件成功: {file_path}，共 {len(data)} 条数据")
        return data

    @staticmethod
    def _load_csv(file_path: Text) -> List[Dict[Text, Any]]:
        """解析 CSV 文件，首行作为表头（字段名）"""
        data = []
        # encoding 用 utf-8-sig 兼容带 BOM 的 CSV
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                data.append(dict(row))
        return data

    @staticmethod
    def _load_json(file_path: Text) -> List[Dict[Text, Any]]:
        """解析 JSON 文件，支持数组格式或 {data: [...]}、单条对象 3种格式"""
        with open(file_path, "r", encoding="utf-8") as f:
            content = json.load(f)

        # 如果是数组，直接返回
        if isinstance(content, list):
            return content
        # 如果是字典且包含 data 键，返回 data
        if isinstance(content, dict) and "data" in content:
            return content["data"]
        # 单条记录也包装成列表返回
        if isinstance(content, dict):
            return [content]
        raise ValueError(f"JSON 文件格式不支持，应为数组或包含 data 键的对象: {file_path}")

    @staticmethod
    def _load_excel(file_path: Text) -> List[Dict[Text, Any]]:
        """解析 Excel 文件，首行作为表头"""
        # 延迟导入，避免未安装 openpyxl 时影响其他格式
        from openpyxl import load_workbook

        data = []
        wb = load_workbook(file_path, read_only=True, data_only=True)
        # 默认取第一个工作表
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return data

        # 首行作为表头
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        # 其余行作为数据
        for row in rows[1:]:
            # 全空行跳过
            if all(cell is None for cell in row):
                continue
            # 按表头组装字典
            row_data = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = row[idx] if idx < len(row) else None
                row_data[header] = value
            data.append(row_data)

        wb.close()
        return data

    @classmethod
    def load_as_params(cls, file_path: Text) -> List[tuple]:
        """加载并转换为 pytest.param 友好的 tuple 列表

        用法:
            @pytest.mark.parametrize("params", DataLoader.load_as_params("data.csv"))
            def test_xxx(params): ...

        Returns:
            List[tuple]: 每条数据包装成单元素 tuple，便于 parametrize 使用
        """
        data_list = cls.load(file_path)
        return [(item,) for item in data_list]
